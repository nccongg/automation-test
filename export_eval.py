#!/usr/bin/env python3
# python export_eval.py --base-url-contains orangehrm
"""
Export evaluation log to Excel/CSV.

Usage:
    pip install psycopg2-binary openpyxl python-dotenv
    python export_eval.py                  # -> eval_log_<timestamp>.xlsx
    python export_eval.py --format csv     # -> eval_log_<timestamp>.csv
    python export_eval.py --project 1      # filter by project id
    python export_eval.py --orangehrm      # all projects whose base_url contains orangehrm
    python export_eval.py --base-url-contains orangehrm

Excel output contains, in order:
    1. Case Study Summary    - headline case-study metrics (Metric / Value)
    2. Execution Mode Summary - Agent vs replay-script execution metrics
    3. Module Summary        - per-module pass/fail/reuse/timing breakdown
    4. Failure Analysis      - failure types ranked by frequency
    5. Tree Evaluation Summary - Generated Tree Correct? breakdown
    6. Evaluation Log        - raw per-run data
    7. TC Generation Time    - raw per-batch data
    8. Dataset Gen Time      - raw per-batch data
    9. Step Failures         - raw per-step data
    10. Column Descriptions  - glossary for every sheet above

CSV output only ever contains the main Evaluation Log.
"""

import argparse
import csv
import os
import sys
from datetime import datetime
from pathlib import Path

try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    sys.exit("Missing dependency: pip install psycopg2-binary")

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent / "server" / ".env")
except ImportError:
    pass


# ── DB ─────────────────────────────────────────────────────────────────────────

def get_conn():
    database_url = os.getenv("DATABASE_URL")
    if database_url:
        return psycopg2.connect(database_url, sslmode="require")
    return psycopg2.connect(
        host=os.getenv("DB_HOST", "localhost"),
        port=int(os.getenv("DB_PORT", 5432)),
        dbname=os.getenv("DB_NAME", "automation_test"),
        user=os.getenv("DB_USER", "postgres"),
        password=os.getenv("DB_PASSWORD", ""),
    )


# ── Queries ────────────────────────────────────────────────────────────────────
# All queries are executed with a params dict (see fetch()), so every literal
# "%" in LIKE patterns below is doubled ("%%") to survive psycopg2's %-style
# parameter substitution. Project/base URL filters are always bound via params
# such as %(project_id)s and %(base_url_pattern)s — never interpolated into SQL.

# Sheet: Bảng đánh giá chính
QUERY_EVAL_LOG = """
SELECT
  'TC' || LPAD(tc.id::text, 2, '0')                          AS "Task ID",

  CASE
    WHEN LOWER(COALESCE(tra.agent_prompt, tc.goal, '')) LIKE '%%login%%'
      OR LOWER(COALESCE(tra.agent_prompt, tc.goal, '')) LIKE '%%sign in%%'
      OR LOWER(COALESCE(tra.agent_prompt, tc.goal, '')) LIKE '%%đăng nhập%%'
      THEN 'Login'
    WHEN LOWER(COALESCE(tra.agent_prompt, tc.goal, '')) LIKE '%%pim%%'
      OR LOWER(COALESCE(tra.agent_prompt, tc.goal, '')) LIKE '%%employee%%'
      OR LOWER(COALESCE(tra.agent_prompt, tc.goal, '')) LIKE '%%nhân viên%%'
      THEN 'PIM'
    WHEN LOWER(COALESCE(tra.agent_prompt, tc.goal, '')) LIKE '%%leave%%'
      OR LOWER(COALESCE(tra.agent_prompt, tc.goal, '')) LIKE '%%nghỉ phép%%'
      THEN 'Leave'
    WHEN LOWER(COALESCE(tra.agent_prompt, tc.goal, '')) LIKE '%%recruit%%'
      OR LOWER(COALESCE(tra.agent_prompt, tc.goal, '')) LIKE '%%tuyển dụng%%'
      THEN 'Recruitment'
    WHEN LOWER(COALESCE(tra.agent_prompt, tc.goal, '')) LIKE '%%time%%'
      OR LOWER(COALESCE(tra.agent_prompt, tc.goal, '')) LIKE '%%attendance%%'
      THEN 'Time & Attendance'
    WHEN LOWER(COALESCE(tra.agent_prompt, tc.goal, '')) LIKE '%%admin%%'
      THEN 'Admin'
    ELSE 'Other'
  END                                                         AS "Module",

  -- Prompt gốc người dùng nhập để sinh test case
  COALESCE(b.source_prompt, '')                              AS "User Prompt",

  -- Test case được sinh ra (goal / agent prompt)
  COALESCE(tra.agent_prompt, tc.goal)                        AS "Generated Test Case",

  COALESCE(tr.generated_tree_correct, '')                    AS "Generated Tree Correct?",

  CASE tr.verdict
    WHEN 'pass'              THEN 'Pass'
    WHEN 'pass_with_warning' THEN 'Pass'
    WHEN 'fail'              THEN 'Fail'
    WHEN 'error'             THEN 'Fail'
    ELSE COALESCE(tr.verdict, '')
  END                                                        AS "Execution Result",

  CASE
    WHEN tr.verdict IN ('pass', 'pass_with_warning') THEN ''
    ELSE COALESCE((
      SELECT
        CASE rsl.failure_reason
          WHEN 'element_not_found'   THEN 'Bad locator'
          WHEN 'element_not_visible' THEN 'Bad locator'
          WHEN 'selector_invalid'    THEN 'Bad locator'
          WHEN 'timeout'             THEN 'Timeout'
          WHEN 'navigation_failed'   THEN 'App error'
          WHEN 'assertion_mismatch'  THEN 'Prompt misunderstanding'
          WHEN 'value_not_set'       THEN 'Prompt misunderstanding'
          WHEN 'unexpected_error'    THEN 'App error'
          ELSE rsl.failure_reason
        END
      FROM run_step_logs rsl
      WHERE rsl.test_run_id = tr.id
        AND rsl.status = 'failed'
        AND rsl.failure_reason IS NOT NULL
      ORDER BY rsl.step_no
      LIMIT 1
    ), '')
  END                                                        AS "Failure Type",

  -- Thời gian LLM sinh test case (từ batch generation)
  CASE
    WHEN b.generation_duration_ms IS NOT NULL
    THEN ROUND(b.generation_duration_ms / 1000.0, 2)::text || 's'
    ELSE ''
  END                                                        AS "Generation Time",

  COALESCE(b.input_tokens::text, '')                         AS "Input Tokens",
  COALESCE(b.output_tokens::text, '')                        AS "Output Tokens",
  CASE
    WHEN b.input_tokens IS NOT NULL AND b.output_tokens IS NOT NULL
    THEN (b.input_tokens + b.output_tokens)::text
    ELSE ''
  END                                                        AS "Total Tokens",

  COALESCE(tr.agent_input_tokens::text, '')                  AS "Agent Input Tokens",
  COALESCE(tr.agent_output_tokens::text, '')                 AS "Agent Output Tokens",
  CASE
    WHEN tr.agent_input_tokens IS NOT NULL AND tr.agent_output_tokens IS NOT NULL
    THEN (tr.agent_input_tokens + tr.agent_output_tokens)::text
    ELSE ''
  END                                                        AS "Agent Total Tokens",

  -- Thời gian thực thi test case trên trình duyệt
  CASE
    WHEN tr.started_at IS NOT NULL AND tr.finished_at IS NOT NULL
    THEN ROUND(EXTRACT(EPOCH FROM (tr.finished_at - tr.started_at))::numeric, 1)::text || 's'
    WHEN tr.execution_log->>'totalDurationSeconds' IS NOT NULL
    THEN tr.execution_log->>'totalDurationSeconds' || 's'
    ELSE ''
  END                                                        AS "Execution Time",

  -- Tổng thời gian = generation + execution
  CASE
    WHEN b.generation_duration_ms IS NOT NULL
      AND tr.started_at IS NOT NULL AND tr.finished_at IS NOT NULL
    THEN ROUND(
      (b.generation_duration_ms / 1000.0
        + EXTRACT(EPOCH FROM (tr.finished_at - tr.started_at))
      )::numeric, 1
    )::text || 's'
    ELSE ''
  END                                                        AS "Total Time",

  CASE
    WHEN tr.verdict = 'pass'
      AND EXISTS (SELECT 1 FROM execution_scripts es WHERE es.source_test_run_id = tr.id) THEN 'Yes'
    WHEN tr.verdict = 'pass_with_warning'
      AND EXISTS (SELECT 1 FROM execution_scripts es WHERE es.source_test_run_id = tr.id) THEN 'Partial'
    WHEN tr.verdict IN ('pass', 'pass_with_warning')            THEN 'Partial'
    ELSE 'No'
  END                                                        AS "Reusable?",

  COALESCE(tr.run_notes,
    CASE
      WHEN tr.verdict IN ('fail','error')
      THEN COALESCE((
        SELECT rsl.message
        FROM run_step_logs rsl
        WHERE rsl.test_run_id = tr.id AND rsl.status = 'failed'
        ORDER BY rsl.step_no LIMIT 1
      ), tr.error_message)
      ELSE NULL
    END
  , '')                                                      AS "Notes",

  tr.id                                                      AS "Run ID",
  tra.trigger_type                                           AS "Run Type",
  p.name                                                     AS "Project",
  p.base_url                                                 AS "Project URL",
  tr.started_at                                              AS "Run Date"

FROM test_runs tr
JOIN test_cases tc   ON tc.id = tr.test_case_id
JOIN projects p      ON p.id  = tc.project_id
-- A test run can have several attempts (retries); take only the latest one so
-- this join can never multiply a test_runs row (which would silently inflate
-- Total Runs and bias any non-deduplicated average, e.g. Execution/Total Time).
LEFT JOIN LATERAL (
  SELECT tra2.agent_prompt, tra2.trigger_type
  FROM test_run_attempts tra2
  WHERE tra2.test_run_id = tr.id
  ORDER BY tra2.attempt_no DESC
  LIMIT 1
) tra ON TRUE
-- A test case can be regenerated more than once, so several generation
-- batches may reference the same selected_test_case_id. Take only the most
-- recent one so this join, too, can never multiply a test_runs row.
LEFT JOIN LATERAL (
  SELECT b2.source_prompt, b2.generation_duration_ms, b2.input_tokens, b2.output_tokens
  FROM test_case_generation_candidates tgc2
  JOIN test_case_generation_batches b2 ON b2.id = tgc2.batch_id
  WHERE tgc2.selected_test_case_id = tc.id
    AND tgc2.is_selected = TRUE
  ORDER BY tgc2.created_at DESC
  LIMIT 1
) b ON TRUE
WHERE tr.status IN ('completed', 'failed')
  {project_filter}
ORDER BY tr.started_at DESC
"""

# Sheet: Thời gian sinh testcase (LLM generation)
QUERY_GENERATION_TIME = """
SELECT
  'TC' || LPAD(tc.id::text, 2, '0')                        AS "Task ID",
  tc.title                                                  AS "Test Case",
  b.source_prompt                                           AS "Prompt",
  b.llm_provider                                           AS "LLM Provider",
  b.llm_model                                              AS "LLM Model",
  b.candidate_count                                        AS "Candidates",
  b.generation_started_at                                  AS "Started At",
  b.generation_finished_at                                 AS "Finished At",
  CASE
    WHEN b.generation_duration_ms IS NOT NULL
    THEN ROUND(b.generation_duration_ms / 1000.0, 2)::text || 's'
    ELSE ''
  END                                                      AS "Duration",
  COALESCE(b.input_tokens::text, '')                       AS "Input Tokens",
  COALESCE(b.output_tokens::text, '')                      AS "Output Tokens",
  CASE
    WHEN b.input_tokens IS NOT NULL AND b.output_tokens IS NOT NULL
    THEN (b.input_tokens + b.output_tokens)::text
    ELSE ''
  END                                                      AS "Total Tokens",
  p.name                                                   AS "Project",
  p.base_url                                               AS "Project URL",
  b.created_at                                             AS "Created At"
FROM test_case_generation_batches b
JOIN projects p ON p.id = b.project_id
LEFT JOIN test_case_generation_candidates tgc ON tgc.batch_id = b.id AND tgc.is_selected = TRUE
LEFT JOIN test_cases tc ON tc.id = tgc.selected_test_case_id
WHERE TRUE
{project_filter}
ORDER BY b.created_at DESC
"""

# Sheet: Thời gian sinh dataset
QUERY_DATASET_GEN_TIME = """
SELECT
  p.name                                                   AS "Project",
  p.base_url                                               AS "Project URL",
  dgl.prompt                                               AS "Prompt",
  dgl.row_count                                            AS "Row Count",
  dgl.llm_provider                                        AS "LLM Provider",
  dgl.llm_model                                           AS "LLM Model",
  dgl.started_at                                          AS "Started At",
  dgl.finished_at                                         AS "Finished At",
  CASE
    WHEN dgl.duration_ms IS NOT NULL
    THEN ROUND(dgl.duration_ms / 1000.0, 2)::text || 's'
    ELSE ''
  END                                                     AS "Duration",
  COALESCE(dgl.input_tokens::text, '')                     AS "Input Tokens",
  COALESCE(dgl.output_tokens::text, '')                    AS "Output Tokens",
  CASE
    WHEN dgl.input_tokens IS NOT NULL AND dgl.output_tokens IS NOT NULL
    THEN (dgl.input_tokens + dgl.output_tokens)::text
    ELSE ''
  END                                                     AS "Total Tokens",
  CASE WHEN dgl.success THEN 'Success' ELSE 'Failed' END  AS "Result",
  COALESCE(dgl.error_message, '')                         AS "Error",
  dgl.created_at                                          AS "Created At"
FROM dataset_generation_logs dgl
LEFT JOIN projects p ON p.id = dgl.project_id
WHERE TRUE
{project_filter}
ORDER BY dgl.created_at DESC
"""

# Sheet: Chi tiết từng bước thất bại
QUERY_STEP_FAILURES = """
SELECT
  'TC' || LPAD(tc.id::text, 2, '0')                       AS "Task ID",
  tc.title                                                 AS "Test Case",
  tr.id                                                    AS "Run ID",
  rsl.step_no                                              AS "Step",
  rsl.action                                               AS "Action",
  CASE rsl.failure_reason
    WHEN 'element_not_found'   THEN 'Bad locator'
    WHEN 'element_not_visible' THEN 'Bad locator'
    WHEN 'selector_invalid'    THEN 'Bad locator'
    WHEN 'timeout'             THEN 'Timeout'
    WHEN 'navigation_failed'   THEN 'App error'
    WHEN 'assertion_mismatch'  THEN 'Prompt misunderstanding'
    WHEN 'value_not_set'       THEN 'Prompt misunderstanding'
    WHEN 'unexpected_error'    THEN 'App error'
    ELSE COALESCE(rsl.failure_reason, '')
  END                                                      AS "Failure Type",
  rsl.message                                              AS "Detail",
  rsl.current_url                                          AS "URL",
  p.name                                                   AS "Project",
  p.base_url                                               AS "Project URL",
  tr.started_at                                            AS "Run Date"
FROM run_step_logs rsl
JOIN test_runs tr  ON tr.id = rsl.test_run_id
JOIN test_cases tc ON tc.id = tr.test_case_id
JOIN projects p    ON p.id  = tc.project_id
WHERE rsl.status = 'failed'
  {project_filter}
ORDER BY tr.started_at DESC, rsl.step_no
"""

QUERY_PROJECT_NAME = 'SELECT name FROM projects WHERE id = %(project_id)s'

QUERY_PROJECTS_BY_BASE_URL = """
SELECT id, name, base_url
FROM projects
WHERE base_url ILIKE %(base_url_pattern)s
ORDER BY name
"""


# ── Helpers ────────────────────────────────────────────────────────────────────

def fetch(conn, sql, params):
    with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
        cur.execute(sql, params)
        rows = cur.fetchall()
        cols = [d[0] for d in cur.description]
    return cols, [list(r) for r in rows]


def clean(val):
    if val is None:
        return ""
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    return val


def clean_row(row):
    return [clean(v) for v in row]


def to_records(cols, rows):
    return [dict(zip(cols, row)) for row in rows]


def build_scope_filter(project_id=None, base_url_contains=None):
    clauses = []
    params = {}

    if project_id is not None:
        params["project_id"] = project_id
        clauses.append("p.id = %(project_id)s")

    if base_url_contains:
        params["base_url_pattern"] = f"%{base_url_contains}%"
        clauses.append("p.base_url ILIKE %(base_url_pattern)s")

    return ("AND " + " AND ".join(clauses)) if clauses else "", params


# ── Numeric parsing (Generation/Execution/Total Time are stored as "8.2s") ─────

def to_seconds(val):
    """Parse a duration value ("8.2s", 8.2, "", None) into float seconds or None."""
    if val in (None, ""):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    text = str(val).strip()
    if text.endswith("s"):
        text = text[:-1]
    try:
        return float(text)
    except ValueError:
        return None


def to_int(val):
    if val in (None, ""):
        return None
    try:
        return int(val)
    except (TypeError, ValueError):
        return None


def avg(values):
    present = [v for v in values if v is not None]
    return sum(present) / len(present) if present else None


def round_or_blank(value, ndigits=2):
    return round(value, ndigits) if value is not None else ""


def sum_or_blank(values, ndigits=2):
    present = [v for v in values if v is not None]
    return round(sum(present), ndigits) if present else ""


def safe_rate(part, whole):
    return (part / whole) if whole else 0.0


def normalize_tree_result(val):
    """DB stores generated_tree_correct as lowercase 'yes'/'partial'/'no' (see
    migration 020, chk_tree_correct). Normalize defensively either way."""
    text = (val or "").strip().lower()
    return text if text in ("yes", "partial", "no") else ""


def is_replay_run(record):
    return (record.get("Run Type") or "").strip().lower() == "manual_replay"


def execution_mode_stats(records):
    passed_records = [r for r in records if r["Execution Result"] == "Pass"]
    failed_records = [r for r in records if r["Execution Result"] == "Fail"]

    exec_seconds = [to_seconds(r["Execution Time"]) for r in records]
    passed_seconds = [to_seconds(r["Execution Time"]) for r in passed_records]
    failed_seconds = [to_seconds(r["Execution Time"]) for r in failed_records]

    return {
        "total_tasks": len({r["Task ID"] for r in records}),
        "total_runs": len(records),
        "passed_tasks": len({r["Task ID"] for r in passed_records}),
        "passed_runs": len(passed_records),
        "failed_tasks": len({r["Task ID"] for r in failed_records}),
        "failed_runs": len(failed_records),
        "pass_rate": Pct(safe_rate(len(passed_records), len(records))),
        "total_time": sum_or_blank(exec_seconds),
        "passed_time": sum_or_blank(passed_seconds),
        "failed_time": sum_or_blank(failed_seconds),
        "avg_time": round_or_blank(avg(exec_seconds)),
        "avg_passed_time": round_or_blank(avg(passed_seconds)),
        "avg_failed_time": round_or_blank(avg(failed_seconds)),
    }


class Pct:
    """Marks a value as a fraction (0..1) to be rendered with Excel's percent format."""
    __slots__ = ("value",)

    def __init__(self, value):
        self.value = value


# ── Excel export ───────────────────────────────────────────────────────────────

def export_excel(sheets: dict, path: str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("Missing dependency: pip install openpyxl")

    wb = Workbook()
    wb.remove(wb.active)

    HDR_FILL   = PatternFill("solid", fgColor="1E3A5F")
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=11)
    ALT_FILL   = PatternFill("solid", fgColor="F0F4FA")
    PASS_FONT  = Font(color="1A7A3C", bold=True)
    FAIL_FONT  = Font(color="C0392B", bold=True)
    THIN       = Side(style="thin", color="D0D7E3")
    BORDER     = Border(bottom=Side(style="thin", color="E8ECF4"))

    def display_str(val):
        if isinstance(val, Pct):
            return f"{val.value * 100:.1f}%"
        return str(clean(val))

    for sheet_name, (cols, rows) in sheets.items():
        ws = wb.create_sheet(title=sheet_name)

        # Header
        for ci, name in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=name)
            c.font      = HDR_FONT
            c.fill      = HDR_FILL
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 32

        # Data
        result_col = next((i+1 for i, n in enumerate(cols) if n == "Execution Result"), None)
        for ri, row in enumerate(rows, 2):
            bg = ALT_FILL if ri % 2 == 0 else None
            for ci, raw_val in enumerate(row, 1):
                if isinstance(raw_val, Pct):
                    c = ws.cell(row=ri, column=ci, value=raw_val.value)
                    c.number_format = "0.0%"
                    val = raw_val.value
                else:
                    val = clean(raw_val)
                    c = ws.cell(row=ri, column=ci, value=val)
                c.alignment = Alignment(vertical="center", wrap_text=False)
                c.border    = BORDER
                if bg:
                    c.fill = bg
                # Tô màu cột Execution Result
                if ci == result_col:
                    if val == "Pass":
                        c.font = PASS_FONT
                    elif val == "Fail":
                        c.font = FAIL_FONT

        # Auto column width
        for ci, name in enumerate(cols, 1):
            vals    = [str(name)] + [display_str(r[ci-1]) for r in rows]
            max_len = min(max((len(v) for v in vals), default=10), 80)
            ws.column_dimensions[get_column_letter(ci)].width = max_len + 3

        ws.freeze_panes = "A2"

    wb.save(path)


# ── CSV export ─────────────────────────────────────────────────────────────────

def export_csv(cols, rows, path: str):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(cols)
        for row in rows:
            w.writerow(clean_row(row))


# ── Summary builders (Python aggregation over the Evaluation Log rows) ─────────

def build_case_study_summary(eval_cols, eval_rows, project_label):
    recs = to_records(eval_cols, eval_rows)
    total_runs = len(recs)
    total_tasks = len({r["Task ID"] for r in recs})

    tree_results = [normalize_tree_result(r["Generated Tree Correct?"]) for r in recs]
    correct   = tree_results.count("yes")
    partial   = tree_results.count("partial")
    incorrect = tree_results.count("no")
    evaluated = correct + partial + incorrect

    passed = sum(1 for r in recs if r["Execution Result"] == "Pass")
    failed = sum(1 for r in recs if r["Execution Result"] == "Fail")

    reusable = sum(1 for r in recs if r["Reusable?"] == "Yes")

    # A "manual_replay" run reuses an already-generated action tree/script —
    # it does not generate a new tree, so it must not be counted as one.
    replay_records = [r for r in recs if is_replay_run(r)]
    agent_records = [r for r in recs if not is_replay_run(r)]
    replay_runs = len(replay_records)
    generated_trees = total_runs - replay_runs
    agent_stats = execution_mode_stats(agent_records)
    replay_stats = execution_mode_stats(replay_records)

    # Generation time / tokens belong to the test case's generation batch, not
    # to each individual run — dedupe by Task ID so replays don't double-count.
    first_by_task = {}
    for r in recs:
        first_by_task.setdefault(r["Task ID"], r)
    gen_seconds   = [to_seconds(r["Generation Time"]) for r in first_by_task.values()]
    gen_input_tok  = [to_int(r["Input Tokens"]) for r in first_by_task.values()]
    gen_output_tok = [to_int(r["Output Tokens"]) for r in first_by_task.values()]

    exec_seconds  = [to_seconds(r["Execution Time"]) for r in recs]
    total_seconds = [to_seconds(r["Total Time"]) for r in recs]
    agent_input_tok  = [to_int(r["Agent Input Tokens"]) for r in recs]
    agent_output_tok = [to_int(r["Agent Output Tokens"]) for r in recs]

    rows = [
        ["Project / Case Study",              project_label],
        ["Total Tasks / Test Cases",          total_tasks],
        ["Total Executions / Runs",           total_runs],
        ["Total Generated Trees",             generated_trees],
        ["Agent Test Cases Run",              agent_stats["total_tasks"]],
        ["Agent Runs",                        agent_stats["total_runs"]],
        ["Agent Total Execution Time (s)",    agent_stats["total_time"]],
        ["Agent Passed Test Cases",           agent_stats["passed_tasks"]],
        ["Agent Passed Runs",                 agent_stats["passed_runs"]],
        ["Agent Passed Execution Time (s)",   agent_stats["passed_time"]],
        ["Agent Failed Test Cases",           agent_stats["failed_tasks"]],
        ["Agent Failed Runs",                 agent_stats["failed_runs"]],
        ["Agent Failed Execution Time (s)",   agent_stats["failed_time"]],
        ["Agent Avg Passed Time (s)",         agent_stats["avg_passed_time"]],
        ["Agent Avg Failed Time (s)",         agent_stats["avg_failed_time"]],
        ["Replay Script Test Cases Run",      replay_stats["total_tasks"]],
        ["Replay Script Runs",                replay_stats["total_runs"]],
        ["Replay Script Total Execution Time (s)", replay_stats["total_time"]],
        ["Replay Script Passed Test Cases",   replay_stats["passed_tasks"]],
        ["Replay Script Passed Runs",         replay_stats["passed_runs"]],
        ["Replay Script Passed Execution Time (s)", replay_stats["passed_time"]],
        ["Replay Script Failed Test Cases",   replay_stats["failed_tasks"]],
        ["Replay Script Failed Runs",         replay_stats["failed_runs"]],
        ["Replay Script Failed Execution Time (s)", replay_stats["failed_time"]],
        ["Replay Script Avg Passed Time (s)", replay_stats["avg_passed_time"]],
        ["Replay Script Avg Failed Time (s)", replay_stats["avg_failed_time"]],
        ["Correct Trees",                     correct],
        ["Partially Correct Trees",           partial],
        ["Incorrect Trees",                   incorrect],
        ["Tree Correctness Rate",             Pct(safe_rate(correct, evaluated))],
        ["Passed Executions",                 passed],
        ["Failed Executions",                 failed],
        ["Execution Pass Rate",               Pct(safe_rate(passed, total_runs))],
        ["Execution Fail Rate",               Pct(safe_rate(failed, total_runs))],
        ["Reusable Count",                    reusable],
        ["Reuse Rate",                        Pct(safe_rate(reusable, total_runs))],
        ["Average Generation Time (s)",       round_or_blank(avg(gen_seconds))],
        ["Average Execution Time (s)",        round_or_blank(avg(exec_seconds))],
        ["Average Total Time (s)",            round_or_blank(avg(total_seconds))],
        ["Total Input Tokens",                sum(t for t in gen_input_tok if t is not None)],
        ["Total Output Tokens",               sum(t for t in gen_output_tok if t is not None)],
        ["Average Input Tokens",              round_or_blank(avg(gen_input_tok))],
        ["Average Output Tokens",             round_or_blank(avg(gen_output_tok))],
        ["Total Agent Input Tokens",          sum(t for t in agent_input_tok if t is not None)],
        ["Total Agent Output Tokens",         sum(t for t in agent_output_tok if t is not None)],
        ["Average Agent Input Tokens",        round_or_blank(avg(agent_input_tok))],
        ["Average Agent Output Tokens",       round_or_blank(avg(agent_output_tok))],
    ]
    return ["Metric", "Value"], rows


def build_execution_mode_summary(eval_cols, eval_rows):
    recs = to_records(eval_cols, eval_rows)
    groups = [
        ("Agent", [r for r in recs if not is_replay_run(r)]),
        ("Replay Script", [r for r in recs if is_replay_run(r)]),
    ]

    cols = [
        "Execution Mode",
        "Total Test Cases",
        "Total Runs",
        "Passed Test Cases",
        "Passed Runs",
        "Failed Test Cases",
        "Failed Runs",
        "Pass Rate",
        "Total Execution Time (s)",
        "Passed Execution Time (s)",
        "Failed Execution Time (s)",
        "Avg Execution Time (s)",
        "Avg Passed Time (s)",
        "Avg Failed Time (s)",
    ]
    rows = []
    for mode, records in groups:
        stats = execution_mode_stats(records)
        rows.append([
            mode,
            stats["total_tasks"],
            stats["total_runs"],
            stats["passed_tasks"],
            stats["passed_runs"],
            stats["failed_tasks"],
            stats["failed_runs"],
            stats["pass_rate"],
            stats["total_time"],
            stats["passed_time"],
            stats["failed_time"],
            stats["avg_time"],
            stats["avg_passed_time"],
            stats["avg_failed_time"],
        ])
    return cols, rows


def build_module_summary(eval_cols, eval_rows):
    recs = to_records(eval_cols, eval_rows)
    by_module = {}
    for r in recs:
        by_module.setdefault(r["Module"], []).append(r)

    cols = [
        "Module", "Total Tasks", "Total Runs", "Pass Count", "Fail Count",
        "Pass Rate", "Fail Rate", "Reusable Count", "Reuse Rate",
        "Avg Generation Time (s)", "Avg Execution Time (s)", "Avg Total Time (s)",
    ]
    rows = []
    for module in sorted(by_module):
        mrecs = by_module[module]
        total_runs  = len(mrecs)
        total_tasks = len({r["Task ID"] for r in mrecs})
        passed   = sum(1 for r in mrecs if r["Execution Result"] == "Pass")
        failed   = sum(1 for r in mrecs if r["Execution Result"] == "Fail")
        reusable = sum(1 for r in mrecs if r["Reusable?"] == "Yes")

        first_by_task = {}
        for r in mrecs:
            first_by_task.setdefault(r["Task ID"], r)
        gen_seconds   = [to_seconds(r["Generation Time"]) for r in first_by_task.values()]
        exec_seconds  = [to_seconds(r["Execution Time"]) for r in mrecs]
        total_seconds = [to_seconds(r["Total Time"]) for r in mrecs]

        rows.append([
            module, total_tasks, total_runs, passed, failed,
            Pct(safe_rate(passed, total_runs)), Pct(safe_rate(failed, total_runs)),
            reusable, Pct(safe_rate(reusable, total_runs)),
            round_or_blank(avg(gen_seconds)),
            round_or_blank(avg(exec_seconds)),
            round_or_blank(avg(total_seconds)),
        ])
    return cols, rows


def build_failure_analysis(eval_cols, eval_rows):
    recs = to_records(eval_cols, eval_rows)
    failed_recs = [r for r in recs if r["Failure Type"]]
    total_failures = len(failed_recs)

    by_type = {}
    for r in failed_recs:
        by_type.setdefault(r["Failure Type"], []).append(r)

    cols = ["Failure Type", "Count", "Percentage", "Related Modules", "Example Failed Task"]
    rows = []
    for ftype in sorted(by_type, key=lambda k: len(by_type[k]), reverse=True):
        frecs = by_type[ftype]
        modules = ", ".join(sorted({r["Module"] for r in frecs}))
        example = frecs[0]
        note = (example["Notes"] or example["Generated Test Case"] or "").strip()
        note = note[:150]
        example_str = f"{example['Task ID']} (Run {example['Run ID']})"
        if note:
            example_str += f": {note}"
        rows.append([ftype, len(frecs), Pct(safe_rate(len(frecs), total_failures)), modules, example_str])
    return cols, rows


def build_tree_evaluation_summary(eval_cols, eval_rows):
    recs = to_records(eval_cols, eval_rows)
    total = len(recs)
    tree_results = [normalize_tree_result(r["Generated Tree Correct?"]) for r in recs]
    yes     = tree_results.count("yes")
    partial = tree_results.count("partial")
    no      = tree_results.count("no")
    empty   = total - yes - partial - no

    cols = ["Result", "Count", "Percentage"]
    rows = [
        ["Yes (Correct)",              yes,     Pct(safe_rate(yes, total))],
        ["Partial (Partially Correct)", partial, Pct(safe_rate(partial, total))],
        ["No (Incorrect)",             no,      Pct(safe_rate(no, total))],
        ["Empty / Not Evaluated",      empty,   Pct(safe_rate(empty, total))],
        ["Total",                      total,   Pct(safe_rate(total, total))],
    ]
    return cols, rows


def distinct_project_names(eval_cols, eval_rows):
    return sorted({r["Project"] for r in to_records(eval_cols, eval_rows) if r.get("Project")})


def resolve_project_label(conn, project_id, base_url_contains, eval_cols, eval_rows):
    if base_url_contains:
        names = distinct_project_names(eval_cols, eval_rows)
        scope = "OrangeHRM" if base_url_contains.lower() == "orangehrm" else (
            f"Projects with base_url containing '{base_url_contains}'"
        )
        if names:
            return f"{scope}: {', '.join(names)}"
        return f"{scope}: N/A"

    if project_id is not None:
        with conn.cursor() as cur:
            cur.execute(QUERY_PROJECT_NAME, {"project_id": project_id})
            row = cur.fetchone()
        return row[0] if row else f"Project #{project_id}"
    names = distinct_project_names(eval_cols, eval_rows)
    return ", ".join(names) if names else "N/A"


def list_projects(conn):
    with conn.cursor() as cur:
        cur.execute("SELECT id, name, base_url FROM projects ORDER BY name")
        return cur.fetchall()


def list_projects_by_base_url(conn, base_url_contains):
    with conn.cursor() as cur:
        cur.execute(QUERY_PROJECTS_BY_BASE_URL, {"base_url_pattern": f"%{base_url_contains}%"})
        return cur.fetchall()


# ── Column Descriptions ────────────────────────────────────────────────────────

COLUMN_DESCRIPTIONS = {
    "Case Study Summary": [
        ("Project / Case Study",       "Tên project/case study web app được đánh giá (lọc theo --project, --orangehrm hoặc --base-url-contains nếu có)."),
        ("Total Tasks / Test Cases",   "Số lượng test case (Task ID) duy nhất xuất hiện trong dữ liệu đánh giá."),
        ("Total Executions / Runs",    "Tổng số lần thực thi (test_runs) đã hoàn tất hoặc thất bại."),
        ("Total Generated Trees",      "Tổng số cây hành động thực sự được agent sinh mới = Total Executions trừ đi các lần replay lại script có sẵn (Run Type = manual_replay), vì replay không sinh cây mới."),
        ("Agent Test Cases Run",       "Số test case duy nhất được chạy bằng agent (Run Type khác manual_replay)."),
        ("Agent Runs",                 "Tổng số lần chạy bằng agent."),
        ("Agent Total Execution Time (s)", "Tổng thời gian thực thi bằng agent trên trình duyệt (giây), không gồm thời gian sinh test case."),
        ("Agent Passed Test Cases",    "Số test case duy nhất có ít nhất một lần chạy bằng agent thành công."),
        ("Agent Passed Runs",          "Số lần chạy bằng agent có Execution Result = Pass."),
        ("Agent Passed Execution Time (s)", "Tổng thời gian của các lần chạy agent thành công."),
        ("Agent Failed Test Cases",    "Số test case duy nhất có ít nhất một lần chạy bằng agent thất bại."),
        ("Agent Failed Runs",          "Số lần chạy bằng agent có Execution Result = Fail."),
        ("Agent Failed Execution Time (s)", "Tổng thời gian của các lần chạy agent thất bại."),
        ("Agent Avg Passed Time (s)",  "Thời gian trung bình của các lần chạy agent thành công."),
        ("Agent Avg Failed Time (s)",  "Thời gian trung bình của các lần chạy agent thất bại."),
        ("Replay Script Test Cases Run", "Số test case duy nhất được chạy bằng replay script (Run Type = manual_replay)."),
        ("Replay Script Runs",         "Tổng số lần chạy bằng replay script."),
        ("Replay Script Total Execution Time (s)", "Tổng thời gian thực thi bằng replay script trên trình duyệt (giây)."),
        ("Replay Script Passed Test Cases", "Số test case duy nhất có ít nhất một lần replay script thành công."),
        ("Replay Script Passed Runs",  "Số lần replay script có Execution Result = Pass."),
        ("Replay Script Passed Execution Time (s)", "Tổng thời gian của các lần replay script thành công."),
        ("Replay Script Failed Test Cases", "Số test case duy nhất có ít nhất một lần replay script thất bại."),
        ("Replay Script Failed Runs",  "Số lần replay script có Execution Result = Fail."),
        ("Replay Script Failed Execution Time (s)", "Tổng thời gian của các lần replay script thất bại."),
        ("Replay Script Avg Passed Time (s)", "Thời gian trung bình của các lần replay script thành công."),
        ("Replay Script Avg Failed Time (s)", "Thời gian trung bình của các lần replay script thất bại."),
        ("Correct Trees",              "Số cây hành động được đánh giá thủ công là đúng hoàn toàn (Yes)."),
        ("Partially Correct Trees",    "Số cây hành động được đánh giá là đúng một phần (Partial)."),
        ("Incorrect Trees",            "Số cây hành động được đánh giá là sai (No)."),
        ("Tree Correctness Rate",      "Correct Trees / (Correct + Partial + Incorrect) — chỉ tính trên các cây đã được đánh giá thủ công."),
        ("Passed Executions",          "Số lần thực thi có Execution Result = Pass."),
        ("Failed Executions",          "Số lần thực thi có Execution Result = Fail."),
        ("Execution Pass Rate",        "Passed Executions / Total Executions."),
        ("Execution Fail Rate",        "Failed Executions / Total Executions."),
        ("Reusable Count",             "Số lần thực thi có Reusable? = Yes (pass và đã có execution script)."),
        ("Reuse Rate",                 "Reusable Count / Total Executions."),
        ("Average Generation Time (s)", "Thời gian sinh test case trung bình (giây), tính theo từng Task ID (không lặp lại khi test case được chạy lại nhiều lần)."),
        ("Average Execution Time (s)", "Thời gian thực thi trung bình trên trình duyệt (giây), tính trên mỗi lần chạy."),
        ("Average Total Time (s)",     "Tổng thời gian trung bình = Generation Time + Execution Time (giây), tính trên mỗi lần chạy."),
        ("Total Input Tokens",         "Tổng token đầu vào dùng để sinh test case (theo Task ID, không lặp lại)."),
        ("Total Output Tokens",        "Tổng token đầu ra khi sinh test case (theo Task ID, không lặp lại)."),
        ("Average Input Tokens",       "Token đầu vào trung bình để sinh một test case (theo Task ID, không lặp lại). Dễ so sánh giữa các case study có số lượng task khác nhau hơn so với Total."),
        ("Average Output Tokens",      "Token đầu ra trung bình khi sinh một test case (theo Task ID, không lặp lại)."),
        ("Total Agent Input Tokens",   "Tổng token đầu vào agent dùng khi thực thi, cộng dồn trên mọi lần chạy."),
        ("Total Agent Output Tokens",  "Tổng token đầu ra agent khi thực thi, cộng dồn trên mọi lần chạy."),
        ("Average Agent Input Tokens", "Token đầu vào agent trung bình trên mỗi lần chạy."),
        ("Average Agent Output Tokens", "Token đầu ra agent trung bình trên mỗi lần chạy."),
    ],
    "Execution Mode Summary": [
        ("Execution Mode",             "Nhóm thực thi: Agent = chạy agent sinh/điều hướng; Replay Script = chạy lại script có sẵn (Run Type = manual_replay)."),
        ("Total Test Cases",           "Số test case duy nhất đã chạy trong nhóm này."),
        ("Total Runs",                 "Tổng số lần chạy trong nhóm này."),
        ("Passed Test Cases",          "Số test case duy nhất có ít nhất một lần chạy thành công trong nhóm này."),
        ("Passed Runs",                "Số lần chạy thành công trong nhóm này."),
        ("Failed Test Cases",          "Số test case duy nhất có ít nhất một lần chạy thất bại trong nhóm này."),
        ("Failed Runs",                "Số lần chạy thất bại trong nhóm này."),
        ("Pass Rate",                  "Passed Runs / Total Runs."),
        ("Total Execution Time (s)",   "Tổng thời gian thực thi trên trình duyệt của nhóm này (giây)."),
        ("Passed Execution Time (s)",  "Tổng thời gian của các lần chạy thành công trong nhóm này."),
        ("Failed Execution Time (s)",  "Tổng thời gian của các lần chạy thất bại trong nhóm này."),
        ("Avg Execution Time (s)",     "Thời gian thực thi trung bình của toàn bộ lần chạy trong nhóm này."),
        ("Avg Passed Time (s)",        "Thời gian thực thi trung bình của các lần chạy thành công."),
        ("Avg Failed Time (s)",        "Thời gian thực thi trung bình của các lần chạy thất bại."),
    ],
    "Module Summary": [
        ("Module",                     "Nhóm chức năng (Login, PIM, Leave, Recruitment, Time & Attendance, Admin, Other)."),
        ("Total Tasks",                "Số test case duy nhất thuộc module này."),
        ("Total Runs",                 "Tổng số lần thực thi thuộc module này."),
        ("Pass Count",                 "Số lần thực thi Pass."),
        ("Fail Count",                 "Số lần thực thi Fail."),
        ("Pass Rate",                  "Pass Count / Total Runs."),
        ("Fail Rate",                  "Fail Count / Total Runs."),
        ("Reusable Count",             "Số lần thực thi có thể tái sử dụng (Reusable? = Yes)."),
        ("Reuse Rate",                 "Reusable Count / Total Runs."),
        ("Avg Generation Time (s)",    "Thời gian sinh test case trung bình của module (giây), theo Task ID."),
        ("Avg Execution Time (s)",     "Thời gian thực thi trung bình của module (giây)."),
        ("Avg Total Time (s)",         "Tổng thời gian trung bình của module (giây)."),
    ],
    "Failure Analysis": [
        ("Failure Type",               "Loại lỗi: Bad locator, Timeout, App error, Prompt misunderstanding, …"),
        ("Count",                      "Số lần xảy ra loại lỗi này."),
        ("Percentage",                 "Count / tổng số lần thực thi thất bại có phân loại lỗi."),
        ("Related Modules",            "Các module có xảy ra loại lỗi này."),
        ("Example Failed Task",        "Ví dụ một Task ID/Run bị lỗi kèm ghi chú, để minh họa trong luận văn."),
    ],
    "Tree Evaluation Summary": [
        ("Result",                     "Kết quả đánh giá cây hành động: Yes (đúng), Partial (đúng một phần), No (sai), Empty (chưa đánh giá)."),
        ("Count",                      "Số lần thực thi thuộc nhóm kết quả này."),
        ("Percentage",                 "Count / Tổng số lần thực thi."),
    ],
    "Evaluation Log": [
        ("Task ID",                  "Mã định danh test case, sinh tự động theo định dạng TC01, TC02, …"),
        ("Module",                   "Nhóm chức năng được kiểm thử, tự phát hiện từ nội dung test case (Login, PIM, Leave, Recruitment, Time & Attendance, Admin, Other)."),
        ("User Prompt",              "Prompt gốc do người dùng nhập vào hệ thống để yêu cầu sinh test case. Đây là đầu vào của LLM."),
        ("Generated Test Case",      "Nội dung test case (goal/steps) do LLM sinh ra từ User Prompt. Đây là đầu vào thực tế được truyền cho agent khi thực thi."),
        ("Generated Tree Correct?",  "Đánh giá thủ công: cây hành động do LLM sinh ra có đúng không? (Yes / No / Partial — để trống nếu chưa đánh giá)."),
        ("Execution Result",         "Kết quả thực thi trên trình duyệt: Pass (thành công) hoặc Fail (thất bại)."),
        ("Failure Type",             "Loại lỗi khi thất bại: Bad locator (không tìm được phần tử), Timeout (quá thời gian chờ), App error (lỗi ứng dụng), Prompt misunderstanding (agent hiểu sai prompt)."),
        ("Generation Time",          "Thời gian LLM sinh ra test case từ User Prompt (tính bằng giây). Không có giá trị nếu test case được tạo thủ công."),
        ("Input Tokens",             "Số token đầu vào (prompt) gửi lên LLM khi sinh test case."),
        ("Output Tokens",            "Số token đầu ra (response) LLM trả về khi sinh test case."),
        ("Total Tokens",             "Tổng Input + Output Tokens. Dùng để ước tính chi phí API và so sánh độ phức tạp giữa các lần sinh."),
        ("Agent Input Tokens",       "Số token đầu vào LLM dùng trong quá trình agent tự động điều hướng trình duyệt."),
        ("Agent Output Tokens",      "Số token đầu ra LLM trong quá trình agent chạy."),
        ("Agent Total Tokens",       "Tổng token agent dùng khi thực thi. Dùng để ước tính chi phí chạy test."),
        ("Execution Time",           "Thời gian agent Selenium thực thi toàn bộ test case trên trình duyệt (tính bằng giây)."),
        ("Total Time",               "Tổng thời gian = Generation Time + Execution Time. Đo toàn bộ chu trình từ lúc nhập prompt đến khi có kết quả kiểm thử."),
        ("Reusable?",                "Test case có thể tái sử dụng không? Yes = pass và đã có execution script; Partial = pass nhưng chưa có script; No = fail."),
        ("Notes",                    "Ghi chú bổ sung: thông báo lỗi đầu tiên khi fail, hoặc ghi chú thủ công của người dùng."),
        ("Run ID",                   "ID nội bộ của lần chạy (test_runs.id), dùng để tra cứu chi tiết trong DB."),
        ("Run Type",                 "Loại trigger: initial (chạy lần đầu), replay (chạy lại), scheduled, …"),
        ("Project",                  "Tên project chứa test case này."),
        ("Project URL",              "Base URL của website test được cấu hình cho project."),
        ("Run Date",                 "Thời điểm bắt đầu thực thi test case (UTC)."),
    ],
    "TC Generation Time": [
        ("Task ID",       "Mã test case được sinh ra từ batch này (TC01, TC02, …)."),
        ("Test Case",     "Tiêu đề test case được chọn từ batch."),
        ("Prompt",        "Prompt gốc người dùng nhập để sinh batch test case."),
        ("LLM Provider",  "Nhà cung cấp LLM được dùng để sinh test case (ví dụ: google)."),
        ("LLM Model",     "Tên model cụ thể (ví dụ: gemini-2.0-flash)."),
        ("Candidates",    "Số lượng test case được sinh ra trong một lần gọi LLM."),
        ("Started At",    "Thời điểm bắt đầu gọi LLM."),
        ("Finished At",   "Thời điểm LLM trả về kết quả."),
        ("Duration",      "Thời gian LLM xử lý (Finished At − Started At), tính bằng giây."),
        ("Input Tokens",  "Số token prompt gửi lên LLM."),
        ("Output Tokens", "Số token response LLM trả về."),
        ("Total Tokens",  "Tổng Input + Output Tokens của batch này."),
        ("Project",       "Tên project."),
        ("Project URL",   "Base URL của website test được cấu hình cho project."),
        ("Created At",    "Thời điểm tạo batch."),
    ],
    "Dataset Gen Time": [
        ("Project",       "Tên project."),
        ("Project URL",   "Base URL của website test được cấu hình cho project."),
        ("Prompt",        "Prompt dùng để sinh dataset."),
        ("Row Count",     "Số dòng dữ liệu được sinh ra."),
        ("LLM Provider",  "Nhà cung cấp LLM."),
        ("LLM Model",     "Tên model."),
        ("Started At",    "Thời điểm bắt đầu sinh dataset."),
        ("Finished At",   "Thời điểm hoàn thành sinh dataset."),
        ("Duration",      "Thời gian sinh dataset (tính bằng giây)."),
        ("Input Tokens",  "Số token prompt gửi lên LLM."),
        ("Output Tokens", "Số token response LLM trả về."),
        ("Total Tokens",  "Tổng Input + Output Tokens."),
        ("Result",        "Kết quả: Success hoặc Failed."),
        ("Error",         "Thông báo lỗi nếu sinh thất bại."),
        ("Created At",    "Thời điểm tạo log."),
    ],
    "Step Failures": [
        ("Task ID",      "Mã test case."),
        ("Test Case",    "Tiêu đề test case."),
        ("Run ID",       "ID lần chạy."),
        ("Step",         "Số thứ tự bước bị lỗi."),
        ("Action",       "Hành động agent thực hiện tại bước lỗi (click, type, navigate, …)."),
        ("Failure Type", "Loại lỗi: Bad locator, Timeout, App error, Prompt misunderstanding."),
        ("Detail",       "Thông báo lỗi chi tiết từ agent."),
        ("URL",          "URL trang web tại thời điểm xảy ra lỗi."),
        ("Project",      "Tên project chứa test case bị lỗi."),
        ("Project URL",  "Base URL của website test được cấu hình cho project."),
        ("Run Date",     "Thời điểm chạy test."),
    ],
}

def build_column_descriptions():
    cols = ["Sheet", "Column", "Description"]
    rows = []
    for sheet, entries in COLUMN_DESCRIPTIONS.items():
        for col, desc in entries:
            rows.append([sheet, col, desc])
    return cols, rows


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Export evaluation log")
    parser.add_argument("--format", choices=["excel", "csv"], default="excel")
    parser.add_argument("--project", type=int, default=None, help="Filter by project ID")
    parser.add_argument(
        "--base-url-contains",
        default=None,
        help="Filter by projects.base_url substring, e.g. orangehrm",
    )
    parser.add_argument(
        "--orangehrm",
        action="store_true",
        help="Shortcut for --base-url-contains orangehrm",
    )
    parser.add_argument("--output", default=None, help="Output file path")
    args = parser.parse_args()

    base_url_contains = (args.base_url_contains or "").strip() or None
    if args.orangehrm:
        base_url_contains = "orangehrm"

    pf, params = build_scope_filter(args.project, base_url_contains)

    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    ext = "xlsx" if args.format == "excel" else "csv"
    if args.output:
        out = args.output
    elif base_url_contains:
        slug = "".join(ch if ch.isalnum() else "_" for ch in base_url_contains.lower()).strip("_")
        out = f"eval_log_{slug or 'base_url'}_{ts}.{ext}"
    else:
        out = f"eval_log_{ts}.{ext}"

    db_url = os.getenv("DATABASE_URL")
    if db_url:
        from urllib.parse import urlparse
        parsed = urlparse(db_url)
        print(f"Connecting to {parsed.hostname} ...")
    else:
        print(f"Connecting to {os.getenv('DB_HOST','localhost')}:{os.getenv('DB_PORT',5432)} ...")
    conn = get_conn()

    try:
        print("Fetching data...")
        if base_url_contains:
            matching_projects = list_projects_by_base_url(conn, base_url_contains)
            print(f"  Project URL filter: base_url contains '{base_url_contains}' "
                  f"({len(matching_projects)} matching projects)")

        eval_cols, eval_rows = fetch(conn, QUERY_EVAL_LOG.format(project_filter=pf), params)

        if args.format == "csv":
            export_csv(eval_cols, eval_rows, out)
            print(f"  Evaluation log: {len(eval_rows)} runs")
            print(f"\n✓ Exported → {out}")
            return

        if args.project is None and not base_url_contains:
            projects_in_data = distinct_project_names(eval_cols, eval_rows)
            if len(projects_in_data) > 1:
                print(f"\n✗ The evaluation data spans {len(projects_in_data)} projects: "
                      f"{', '.join(projects_in_data)}")
                print("  A thesis case study should evaluate exactly one app. Re-run with --project <id>:\n")
                for pid, pname, base_url in list_projects(conn):
                    print(f"    --project {pid}   {pname} ({base_url})")
                print("\n  To aggregate one app across multiple projects, use "
                      "--base-url-contains <text> (for example: --orangehrm).")
                sys.exit(1)

        gen_cols,  gen_rows  = fetch(conn, QUERY_GENERATION_TIME.format(project_filter=pf), params)
        dgl_cols,  dgl_rows  = fetch(conn, QUERY_DATASET_GEN_TIME.format(project_filter=pf), params)
        fail_cols, fail_rows = fetch(conn, QUERY_STEP_FAILURES.format(project_filter=pf), params)

        print(f"  Evaluation log:     {len(eval_rows)} runs")
        print(f"  TC generation:      {len(gen_rows)} batches")
        print(f"  Dataset generation: {len(dgl_rows)} logs")
        print(f"  Step failures:      {len(fail_rows)} failed steps")

        project_label = resolve_project_label(conn, args.project, base_url_contains, eval_cols, eval_rows)
        case_study_cols, case_study_rows = build_case_study_summary(eval_cols, eval_rows, project_label)
        mode_cols,       mode_rows       = build_execution_mode_summary(eval_cols, eval_rows)
        module_cols,     module_rows     = build_module_summary(eval_cols, eval_rows)
        failure_cols,    failure_rows    = build_failure_analysis(eval_cols, eval_rows)
        tree_cols,       tree_rows       = build_tree_evaluation_summary(eval_cols, eval_rows)
        desc_cols,       desc_rows       = build_column_descriptions()

        export_excel({
            "Case Study Summary":     (case_study_cols, case_study_rows),
            "Execution Mode Summary": (mode_cols,       mode_rows),
            "Module Summary":         (module_cols,     module_rows),
            "Failure Analysis":       (failure_cols,    failure_rows),
            "Tree Evaluation Summary": (tree_cols,       tree_rows),
            "Evaluation Log":         (eval_cols,  eval_rows),
            "TC Generation Time":     (gen_cols,   gen_rows),
            "Dataset Gen Time":       (dgl_cols,   dgl_rows),
            "Step Failures":          (fail_cols,  fail_rows),
            "Column Descriptions":    (desc_cols,  desc_rows),
        }, out)

        print(f"\n✓ Exported → {out}")

    finally:
        conn.close()


if __name__ == "__main__":
    main()
